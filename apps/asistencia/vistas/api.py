from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Max, Func, DateField
from django.utils import timezone
from datetime import date, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings

from ..models import RegistroAcceso

_intentos_fallidos = defaultdict(list)
_MAX_INTENTOS = 3
_BLOQUEO_SEGUNDOS = 900
from ..servicios import procesar_archivo_dat, emparejar_empleados, calcular_reporte, obtener_empleados_sin_registro
from apps.empleados.models import Empleado


def registros_recientes(request):
    hoy_local = timezone.localtime(timezone.now()).date()
    inicio_hoy = timezone.make_aware(datetime.combine(hoy_local, datetime.min.time()))
    fin_hoy = timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))

    registros = RegistroAcceso.objects.filter(
        marcado_en__gte=inicio_hoy,
        marcado_en__lte=fin_hoy,
    ).select_related('empleado').order_by('-marcado_en')[:20]

    datos = []
    for r in registros:
        tz = timezone.localtime(r.marcado_en)
        datos.append({
            'id': r.empleado.id_visual,
            'empleado_pk': r.empleado.pk,
            'nombre': r.empleado.nombre,
            'hora': tz.strftime('%H:%M:%S'),
            'fecha': tz.strftime('%d/%m/%Y'),
            'dispositivo': r.serial_dispositivo,
        })

    return JsonResponse({'registros': datos, 'total': len(datos)})


def _aplicar_filtros(resultados, nombre, id_busqueda):
    filtrados = resultados
    if nombre:
        nombre = nombre.strip().lower()
        filtrados = [r for r in filtrados if nombre in r['nombre'].lower()]
    if id_busqueda:
        id_busqueda = id_busqueda.strip().lower()
        filtrados = [r for r in filtrados if id_busqueda in r['empleado_id'].lower()]
    return filtrados


def _obtener_registros(fecha, nombre, id_busqueda, desde=None, hasta=None):
    if desde and hasta:
        inicio = timezone.make_aware(datetime.combine(desde, datetime.min.time()))
        fin = timezone.make_aware(datetime.combine(hasta, datetime.max.time()))
        qs = RegistroAcceso.objects.filter(
            marcado_en__gte=inicio, marcado_en__lte=fin
        ).select_related('empleado')
        if nombre or id_busqueda:
            pass
        return list(qs)
    if nombre or id_busqueda:
        qs = RegistroAcceso.objects.all().select_related('empleado')
        if fecha:
            inicio = timezone.make_aware(datetime.combine(fecha, datetime.min.time()))
            fin = timezone.make_aware(datetime.combine(fecha, datetime.max.time()))
            qs = qs.filter(marcado_en__gte=inicio, marcado_en__lte=fin)
        return list(qs)
    if fecha:
        inicio = timezone.make_aware(datetime.combine(fecha, datetime.min.time()))
        fin = timezone.make_aware(datetime.combine(fecha, datetime.max.time()))
        return list(RegistroAcceso.objects.filter(
            marcado_en__gte=inicio, marcado_en__lte=fin
        ).select_related('empleado'))
    return []


def reporte_diario(request):
    cadena_fecha = request.GET.get('fecha')
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    nombre = request.GET.get('nombre', '')
    id_busqueda = request.GET.get('id', '')

    fecha_filtro = None
    desde = None
    hasta = None
    if desde_str and hasta_str:
        try:
            desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
    elif cadena_fecha:
        try:
            fecha_filtro = datetime.strptime(cadena_fecha, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Fecha inválida'}, status=400)

    registros = _obtener_registros(fecha_filtro, nombre, id_busqueda, desde, hasta)
    datos = [(r.empleado, r.marcado_en, r.datos_originales) for r in registros]
    from ..servicios import calcular_reporte
    resultados, totales = calcular_reporte(datos)

    if nombre or id_busqueda:
        empleados_filtro = set()
        if nombre:
            n = nombre.strip().lower()
            empleados_filtro.update(
                Empleado.objects.filter(nombre__icontains=n).values_list('pk', flat=True)
            )
        if id_busqueda:
            ib = id_busqueda.strip().lower()
            for e in Empleado.objects.all():
                if ib in e.id_visual.lower():
                    empleados_filtro.add(e.pk)
        resultados = [r for r in resultados if r['empleado_pk'] in empleados_filtro]

    if not fecha_filtro:
        fecha_filtro = date.today()

    return JsonResponse({
        'fecha': fecha_filtro.isoformat(),
        'resultados': resultados,
        'totales': totales,
        'filtros': {'nombre': nombre, 'id': id_busqueda},
    })


def _estilo_excel():
    negro = '000000'
    blanco = 'FFFFFF'
    azul = '1F4E79'
    gris = 'D6E4F0'
    verde = '2E75B6'

    borde = Border(
        left=Side(style='thin', color=negro),
        right=Side(style='thin', color=negro),
        top=Side(style='thin', color=negro),
        bottom=Side(style='thin', color=negro),
    )

    encabezado_font = Font(name='Calibri', bold=True, color=blanco, size=11)
    encabezado_fill = PatternFill(start_color=azul, end_color=azul, fill_type='solid')
    encabezado_align = Alignment(horizontal='center', vertical='center')

    cuerpo_font = Font(name='Calibri', size=11)
    cuerpo_align = Alignment(horizontal='center', vertical='center')
    nombre_align = Alignment(horizontal='left', vertical='center')

    fila_par_fill = PatternFill(start_color=gris, end_color=gris, fill_type='solid')

    return {
        'borde': borde,
        'encabezado_font': encabezado_font,
        'encabezado_fill': encabezado_fill,
        'encabezado_align': encabezado_align,
        'cuerpo_font': cuerpo_font,
        'cuerpo_align': cuerpo_align,
        'nombre_align': nombre_align,
        'fila_par_fill': fila_par_fill,
        'verde_fill': PatternFill(start_color=verde, end_color=verde, fill_type='solid'),
    }


def _escribir_titulo(ws, titulo, num_columnas):
    est = _estilo_excel()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columnas)
    celda = ws.cell(row=1, column=1, value=titulo)
    celda.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    celda.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    celda.alignment = Alignment(horizontal='center', vertical='center')
    celda.border = est['borde']
    ws.row_dimensions[1].height = 30


def _escribir_excel(worksheet, resultados, columnas, titulo=None):
    ws = worksheet
    est = _estilo_excel()
    inicio = 3 if titulo else 1

    if titulo:
        _escribir_titulo(ws, titulo, len(columnas))
        ws.row_dimensions[2].height = 8

    for i, col_name in enumerate(columnas, 1):
        celda = ws.cell(row=inicio, column=i, value=col_name)
        celda.font = est['encabezado_font']
        celda.fill = est['encabezado_fill']
        celda.alignment = est['encabezado_align']
        celda.border = est['borde']

    for idx, fila in enumerate(resultados):
        row_num = idx + inicio + 1
        for col_idx, key in enumerate(columnas, 1):
            valor = fila.get(key, '')
            if isinstance(valor, float):
                celda = ws.cell(row=row_num, column=col_idx, value=valor)
                celda.number_format = '0.00'
            else:
                celda = ws.cell(row=row_num, column=col_idx, value=valor)
            celda.font = est['cuerpo_font']
            celda.border = est['borde']
            al = est['nombre_align'] if key == 'Nombre' else est['cuerpo_align']
            celda.alignment = al
            if idx % 2 == 1:
                celda.fill = est['fila_par_fill']

    for i in range(1, len(columnas) + 1):
        ancho = max(14, min(40, len(columnas[i - 1]) + 4))
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ult_fila = inicio + len(resultados)
    if ult_fila > inicio:
        ws.auto_filter.ref = f"A{inicio}:{get_column_letter(len(columnas))}{ult_fila}"
    ws.freeze_panes = f"A{inicio + 1}"


def _escribir_excel_general(worksheet, grupos_empleados, columnas, titulo=None):
    ws = worksheet
    est = _estilo_excel()
    inicio = 3 if titulo else 1
    fila = inicio

    verde_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    verde_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)

    if titulo:
        _escribir_titulo(ws, titulo, len(columnas))
        ws.row_dimensions[2].height = 8
        fila = 3

    for i, col_name in enumerate(columnas, 1):
        celda = ws.cell(row=fila, column=i, value=col_name)
        celda.font = est['encabezado_font']
        celda.fill = est['encabezado_fill']
        celda.alignment = est['encabezado_align']
        celda.border = est['borde']
    fila += 1

    for g_idx, grupo in enumerate(grupos_empleados):
        etiqueta = f"{grupo['empleado_id']} — {grupo['nombre']}  ({len(grupo['registros'])} registro(s))"
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(columnas))
        celda = ws.cell(row=fila, column=1, value=etiqueta)
        celda.font = verde_font
        celda.fill = verde_fill
        celda.alignment = Alignment(horizontal='left', vertical='center')
        celda.border = est['borde']
        for c in range(1, len(columnas) + 1):
            ws.cell(row=fila, column=c).border = est['borde']
            ws.cell(row=fila, column=c).fill = verde_fill
        ws.row_dimensions[fila].height = 22
        fila += 1

        for r_idx, r in enumerate(grupo['registros']):
            valores = {
                'ID': r['empleado_id'],
                'Nombre': r['nombre'],
                'Fecha': r['fecha'],
                'Comida inicio': r['comida_inicio'],
                'Comida fin': r['comida_fin'],
                'Comida hrs': r['comida_horas'],
                'Jornada inicio': r['jornada_inicio'],
                'Jornada fin': r['jornada_fin'],
                'Jornada hrs': r['jornada_horas'] if r['jornada_horas'] is not None else ('En curso' if r['comida_fin'] else ''),
                'Incidencia': r.get('incidencia', ''),
            }
            for col_idx, key in enumerate(columnas, 1):
                valor = valores.get(key, '')
                if isinstance(valor, float):
                    celda = ws.cell(row=fila, column=col_idx, value=valor)
                    celda.number_format = '0.00'
                else:
                    celda = ws.cell(row=fila, column=col_idx, value=valor)
                celda.font = est['cuerpo_font']
                celda.border = est['borde']
                al = est['nombre_align'] if key == 'Nombre' else est['cuerpo_align']
                celda.alignment = al
                if r_idx % 2 == 1:
                    celda.fill = est['fila_par_fill']
            fila += 1

        fila += 1

    for i in range(1, len(columnas) + 1):
        ancho = max(14, min(40, len(columnas[i - 1]) + 4))
        ws.column_dimensions[get_column_letter(i)].width = ancho

    if fila > inicio:
        ws.auto_filter.ref = f"A{inicio}:{get_column_letter(len(columnas))}{fila - 1}"
    ws.freeze_panes = f"A{inicio + 1}"


def _escribir_excel_individual(worksheet, datos_detalle, empleado, titulo=None):
    ws = worksheet
    est = _estilo_excel()
    inicio = 3 if titulo else 1
    fila = inicio

    azul_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    blanco_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    info_font = Font(name='Calibri', bold=True, size=12)

    num_cols = 8
    if titulo:
        _escribir_titulo(ws, "Reporte", num_cols)
        ws.row_dimensions[2].height = 8
        fila = 3

    if empleado:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=num_cols)
        celda = ws.cell(row=fila, column=1, value=f"{empleado.id_visual} — {empleado.nombre}")
        celda.font = info_font
        celda.fill = azul_fill
        celda.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
        celda.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, num_cols + 1):
            ws.cell(row=fila, column=c).border = est['borde']
            ws.cell(row=fila, column=c).fill = azul_fill
        ws.row_dimensions[fila].height = 24
        fila += 1

    columnas = ['Fecha', 'Comida inicio', 'Comida fin', 'Comida hrs', 'Jornada inicio', 'Jornada fin', 'Jornada hrs', 'Incidencia']
    num_cols = len(columnas)
    for i, col_name in enumerate(columnas, 1):
        celda = ws.cell(row=fila, column=i, value=col_name)
        celda.font = est['encabezado_font']
        celda.fill = est['encabezado_fill']
        celda.alignment = est['encabezado_align']
        celda.border = est['borde']
    fila += 1

    for idx, fila_data in enumerate(datos_detalle):
        valores = [
            fila_data['Fecha'],
            fila_data.get('comida_inicio', ''),
            fila_data.get('comida_fin', ''),
            fila_data.get('comida_horas', ''),
            fila_data.get('jornada_inicio', ''),
            fila_data.get('jornada_fin', ''),
            fila_data.get('jornada_horas', ''),
            fila_data.get('incidencia', ''),
        ]
        for col_idx, valor in enumerate(valores, 1):
            if isinstance(valor, float):
                celda = ws.cell(row=fila, column=col_idx, value=valor)
                celda.number_format = '0.00'
            else:
                celda = ws.cell(row=fila, column=col_idx, value=valor)
            celda.font = est['cuerpo_font']
            celda.border = est['borde']
            celda.alignment = est['cuerpo_align']
            if idx % 2 == 1:
                celda.fill = est['fila_par_fill']
        fila += 1

    for i in range(1, num_cols + 1):
        ancho = max(14, min(40, 14))
        ws.column_dimensions[get_column_letter(i)].width = ancho + 2

    if fila > inicio:
        ws.auto_filter.ref = f"A{inicio}:{get_column_letter(num_cols)}{fila - 1}"
    ws.freeze_panes = f"A{inicio + 1}"


def exportar_reporte_excel(request):
    cadena_fecha = request.GET.get('fecha')
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    empleado_pk = request.GET.get('empleado_pk')
    nombre = request.GET.get('nombre', '')
    id_busqueda = request.GET.get('id', '')

    fecha_filtro = None
    if desde_str and hasta_str:
        try:
            d_desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            d_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
        inicio_dia = timezone.make_aware(datetime.combine(d_desde, datetime.min.time()))
        fin_dia = timezone.make_aware(datetime.combine(d_hasta, datetime.max.time()))
        rango = f"{d_desde.isoformat()}_a_{d_hasta.isoformat()}"
        fecha_filtro = d_hasta
    elif cadena_fecha:
        try:
            fecha_filtro = datetime.strptime(cadena_fecha, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Fecha inválida'}, status=400)
        inicio_dia = timezone.make_aware(datetime.combine(fecha_filtro, datetime.min.time()))
        fin_dia = timezone.make_aware(datetime.combine(fecha_filtro, datetime.max.time()))
        rango = fecha_filtro.isoformat()

    qs = RegistroAcceso.objects.all().select_related('empleado')
    if empleado_pk:
        qs = qs.filter(empleado__pk=empleado_pk)
    if fecha_filtro and not (nombre or id_busqueda):
        qs = qs.filter(marcado_en__gte=inicio_dia, marcado_en__lte=fin_dia)
    elif fecha_filtro and (nombre or id_busqueda):
        qs = qs.filter(marcado_en__gte=inicio_dia, marcado_en__lte=fin_dia)

    registros = list(qs)
    datos = [(r.empleado, r.marcado_en, r.datos_originales) for r in registros]
    from ..servicios import calcular_reporte
    resultados, totales = calcular_reporte(datos)

    if nombre or id_busqueda:
        empleados_filtro = set()
        if nombre:
            n = nombre.strip().lower()
            empleados_filtro.update(
                Empleado.objects.filter(nombre__icontains=n).values_list('pk', flat=True)
            )
        if id_busqueda:
            ib = id_busqueda.strip().lower()
            for e in Empleado.objects.all():
                if ib in e.id_visual.lower():
                    empleados_filtro.add(e.pk)
        resultados = [r for r in resultados if r['empleado_pk'] in empleados_filtro]

    wb = Workbook()

    letrero_fecha = f"del {rango}" if fecha_filtro else "TODOS LOS REGISTROS"
    nombre_emp = ""
    if nombre or id_busqueda:
        nombre_emp = f" - Filtro: {nombre or ''} {id_busqueda or ''}".strip()

    if empleado_pk:
        titulo = "Reporte"
    else:
        titulo = f"Reporte de Asistencia {letrero_fecha}{nombre_emp}"

    columnas_detalle = ['ID', 'Nombre', 'Fecha', 'Comida inicio', 'Comida fin', 'Comida hrs', 'Jornada inicio', 'Jornada fin', 'Jornada hrs', 'Incidencia']
    datos_detalle = []
    for r in resultados:
        datos_detalle.append({
            'ID': r['empleado_id'],
            'Nombre': r['nombre'],
            'Fecha': r['fecha'],
            'comida_inicio': r['comida_inicio'],
            'comida_fin': r['comida_fin'],
            'comida_horas': r['comida_horas'],
            'jornada_inicio': r['jornada_inicio'],
            'jornada_fin': r['jornada_fin'],
            'jornada_horas': r['jornada_horas'] if r['jornada_horas'] is not None else ('En curso' if r['comida_fin'] else ''),
            'incidencia': r.get('incidencia', ''),
        })

    ws1 = wb.active
    ws1.title = 'Detalle'

    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    empleado_pk = request.GET.get('empleado_pk')
    es_general = desde_str and hasta_str and not empleado_pk

    if empleado_pk:
        emp = Empleado.objects.get(pk=empleado_pk) if empleado_pk else None
        _escribir_excel_individual(ws1, datos_detalle, emp, titulo)
    elif es_general:
        agrupados = {}
        for r in resultados:
            key = r['empleado_pk']
            if key not in agrupados:
                agrupados[key] = {
                    'empleado_id': r['empleado_id'],
                    'nombre': r['nombre'],
                    'total_horas': 0,
                    'registros': [],
                }
            agrupados[key]['registros'].append(r)
            if r['jornada_horas'] is not None:
                agrupados[key]['total_horas'] += r['jornada_horas']
            if r['comida_horas'] is not None:
                agrupados[key]['total_horas'] += r['comida_horas']
        for k in agrupados:
            agrupados[k]['registros'].sort(key=lambda x: x['fecha'])
        items = sorted(agrupados.values(), key=lambda x: x['nombre'])
        _escribir_excel_general(ws1, items, columnas_detalle, titulo)
    else:
        _escribir_excel(ws1, datos_detalle, columnas_detalle, titulo)

    if not empleado_pk and totales:
        ws2 = wb.create_sheet('Resumen')
        cols_resumen = ['ID', 'Nombre', 'Días trabajados', 'Total comida', 'Total jornada']
        datos_resumen = []
        for t in totales:
            datos_resumen.append({
                'ID': t['empleado_id'],
                'Nombre': t['nombre'],
                'Días trabajados': t['dias'],
                'Total comida': t['total_comida'],
                'Total jornada': t['total_jornada'],
            })
        _escribir_excel(ws2, datos_resumen, cols_resumen, titulo)

    nombre_archivo = f"reporte_{rango if fecha_filtro else 'todos'}"
    if empleado_pk:
        emp = get_object_or_404(Empleado, pk=empleado_pk)
        nombre_archivo += f"_{emp.nombre.replace(' ', '_')}"
    nombre_archivo += '.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


def estado_sincronizacion(request):
    ultimo_registro = RegistroAcceso.objects.aggregate(ultimo=Max('marcado_en'))['ultimo']
    ultima_sincro = RegistroAcceso.objects.aggregate(ultima=Max('creado_en'))['ultima']
    ahora = timezone.localtime()
    hoy_local = ahora.date()

    inicio_hoy = timezone.make_aware(datetime.combine(hoy_local, datetime.min.time()))
    fin_hoy = timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))

    return JsonResponse({
        'ip_dispositivo': '10.10.0.3',
        'puerto_dispositivo': 4370,
        'ultimo_registro': ultimo_registro.isoformat() if ultimo_registro else None,
        'ultima_sincronizacion': ultima_sincro.isoformat() if ultima_sincro else None,
        'hora_servidor': ahora.isoformat(),
        'registros_hoy': RegistroAcceso.objects.filter(
            marcado_en__gte=inicio_hoy, marcado_en__lte=fin_hoy
        ).count(),
        'total_registros': RegistroAcceso.objects.count(),
        'total_empleados': Empleado.objects.count(),
    })


def lista_empleados(request):
    datos = []
    for e in Empleado.objects.all():
        datos.append({
            'id': e.id,
            'id_visual': e.id_visual,
            'nombre': e.nombre,
        })
    return JsonResponse({'empleados': datos})


def detalle_empleado_api(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)

    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')

    try:
        if fecha_desde:
            d_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        else:
            d_desde = date.today()
        if fecha_hasta:
            d_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            d_hasta = date.today()
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)

    desde_dt = timezone.make_aware(datetime.combine(d_desde, datetime.min.time()))
    hasta_dt = timezone.make_aware(datetime.combine(d_hasta, datetime.max.time()))

    registros = RegistroAcceso.objects.filter(
        empleado=empleado,
        marcado_en__gte=desde_dt,
        marcado_en__lte=hasta_dt,
    ).order_by('marcado_en').select_related('empleado')

    datos = [(r.empleado, r.marcado_en, r.datos_originales) for r in registros]
    resultados, _ = calcular_reporte(datos)

    total_comida = 0
    total_jornada = 0
    for r in resultados:
        if r['comida_horas'] is not None:
            total_comida += r['comida_horas']
        if r['jornada_horas'] is not None:
            total_jornada += r['jornada_horas']

    dias = []
    for r in resultados:
        dias.append({
            'fecha': r['fecha'],
            'comida_inicio': r['comida_inicio'],
            'comida_fin': r['comida_fin'],
            'comida_horas': r['comida_horas'],
            'jornada_inicio': r['jornada_inicio'],
            'jornada_fin': r['jornada_fin'],
            'jornada_horas': r['jornada_horas'],
            'incidencia': r.get('incidencia', ''),
        })

    total_horas = round(total_comida + total_jornada, 2)
    promedio_horas = round(total_horas / len(dias), 2) if dias else 0

    return JsonResponse({
        'empleado': {
            'id': empleado.id,
            'id_visual': empleado.id_visual,
            'nombre': empleado.nombre,
        },
        'fecha_desde': d_desde.isoformat(),
        'fecha_hasta': d_hasta.isoformat(),
        'dias': dias,
        'resumen': {
            'total_dias': len(dias),
            'total_comida': round(total_comida, 2),
            'total_jornada': round(total_jornada, 2),
            'total_horas': total_horas,
            'promedio_horas': promedio_horas,
        },
    })


def registros_por_fecha(request):
    cadena_fecha = request.GET.get('fecha')
    if not cadena_fecha:
        return JsonResponse({'error': 'Parámetro fecha requerido (YYYY-MM-DD)'}, status=400)
    try:
        fecha_filtro = datetime.strptime(cadena_fecha, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Formato inválido. Use YYYY-MM-DD.'}, status=400)

    inicio_dia = timezone.make_aware(datetime.combine(fecha_filtro, datetime.min.time()))
    fin_dia = timezone.make_aware(datetime.combine(fecha_filtro, datetime.max.time()))

    registros = RegistroAcceso.objects.filter(
        marcado_en__gte=inicio_dia,
        marcado_en__lte=fin_dia,
    ).select_related('empleado').order_by('marcado_en')

    datos = []
    for r in registros:
        tz = timezone.localtime(r.marcado_en)
        datos.append({
            'id': r.empleado.id_visual,
            'empleado_pk': r.empleado.pk,
            'nombre': r.empleado.nombre,
            'hora': tz.strftime('%H:%M:%S'),
            'fecha': tz.strftime('%d/%m/%Y'),
            'dispositivo': r.serial_dispositivo,
        })
    return JsonResponse({
        'registros': datos,
        'total': len(datos),
        'fecha': fecha_filtro.isoformat()
    })


def sincronizar_dispositivo(request):
    from ..extraccion import sincronizar_asistencia, sincronizar_empleados

    resultado_empleados = sincronizar_empleados()
    resultado_asistencia = sincronizar_asistencia()

    return JsonResponse({
        'empleados': resultado_empleados,
        'asistencia': resultado_asistencia,
    })


def fechas_disponibles(request):
    fechas = (
        RegistroAcceso.objects
        .annotate(d=Func('marcado_en', function='DATE', output_field=DateField()))
        .values_list('d', flat=True)
        .distinct()
        .order_by('-d')
    )
    return JsonResponse({'fechas': [f.isoformat() for f in fechas if f is not None]})


@csrf_exempt
def subir_archivo(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Usa POST'}, status=405)

    archivo = request.FILES.get('file')
    if not archivo:
        return JsonResponse({'error': 'No se recibi\u00f3 archivo'}, status=400)

    contenido = archivo.read().decode('utf-8', errors='replace')
    registros_crudos = procesar_archivo_dat(contenido)
    emparejados = emparejar_empleados(registros_crudos)

    if not emparejados:
        return JsonResponse({'error': 'No se encontraron registros v\u00e1lidos'}, status=400)

    for emp, dt, linea in emparejados:
        RegistroAcceso.objects.get_or_create(
            empleado=emp,
            marcado_en=dt,
            defaults={'datos_originales': linea, 'serial_dispositivo': 'upload'}
        )

    resultados, totales = calcular_reporte(emparejados)
    sin_registro = obtener_empleados_sin_registro(registros_crudos)

    return JsonResponse({
        'resultados': resultados,
        'totales': totales,
        'sin_registro': sin_registro,
        'total_empleados': len(totales),
        'total_registros': len(emparejados),
    })


@csrf_exempt
def limpiar_registros(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Usa POST'}, status=405)

    ip = request.META.get('REMOTE_ADDR', 'unknown')
    ahora = timezone.now()
    _intentos_fallidos[ip] = [t for t in _intentos_fallidos[ip] if (ahora - t).total_seconds() < _BLOQUEO_SEGUNDOS]

    if len(_intentos_fallidos[ip]) >= _MAX_INTENTOS:
        return JsonResponse({
            'error': 'Demasiados intentos. Espera 15 minutos.'
        }, status=429)

    import json
    try:
        data = json.loads(request.body)
        password = data.get('password', '')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body inválido'}, status=400)

    if password != settings.CLEANUP_PASSWORD:
        _intentos_fallidos[ip].append(ahora)
        restantes = _MAX_INTENTOS - len(_intentos_fallidos[ip])
        return JsonResponse({
            'error': f'Contraseña incorrecta. {restantes} intento(s) restante(s).'
        }, status=403)

    _intentos_fallidos[ip].clear()
    eliminados, _ = RegistroAcceso.objects.all().delete()

    return JsonResponse({
        'ok': True,
        'eliminados': eliminados,
        'mensaje': f'Se eliminaron {eliminados} registros de asistencia'
    })
